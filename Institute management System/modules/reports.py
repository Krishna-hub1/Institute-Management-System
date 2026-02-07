import customtkinter as ctk
from tkinter import messagebox, filedialog
from datetime import datetime
from tkcalendar import DateEntry

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from config import COLORS


class ReportsModule:
    def __init__(self, app):
        self.app = app

    # ================= MAIN VIEW =================

    def show(self):
        for w in self.app.main_content.winfo_children():
            w.destroy()

        header = ctk.CTkFrame(self.app.main_content, fg_color="transparent", height=80)
        header.pack(fill="x", padx=30, pady=(20, 10))

        ctk.CTkLabel(
            header,
            text="📈 Reports & Export",
            font=ctk.CTkFont(size=28, weight="bold")
        ).pack(side="left")

        body = ctk.CTkFrame(self.app.main_content)
        body.pack(fill="both", expand=True, padx=30, pady=20)

        ctk.CTkLabel(
            body,
            text="📊 Export Data to Excel",
            font=ctk.CTkFont(size=22, weight="bold")
        ).pack(pady=30)

        btns = ctk.CTkFrame(body, fg_color="transparent")
        btns.pack(pady=20)

        ctk.CTkButton(
            btns,
            text="📥 Export Students",
            width=220,
            height=55,
            fg_color=COLORS["info"],
            command=self.export_students
        ).pack(pady=10)

        ctk.CTkButton(
            btns,
            text="📥 Export Courses",
            width=220,
            height=55,
            fg_color=COLORS["success"],
            command=self.export_courses
        ).pack(pady=10)

        ctk.CTkButton(
            btns,
            text="📥 Export Attendance",
            width=220,
            height=55,
            fg_color=COLORS["warning"],
            command=self.export_attendance_dialog
        ).pack(pady=10)

    # ================= STUDENTS =================

    def export_students(self):
        students = self.app.db.get_all_students()
        if not students:
            messagebox.showwarning("No Data", "No students to export")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"students_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"

        headers = [
            "Student ID", "First Name", "Last Name", "Gender",
            "DOB", "Email", "Phone", "Address",
            "Course", "Admission Date", "Status"
        ]

        self._write_headers(ws, headers, "1f538d")

        for r, s in enumerate(students, 2):
            ws.cell(r, 1, s["student_id"])
            ws.cell(r, 2, s["first_name"])
            ws.cell(r, 3, s["last_name"])
            ws.cell(r, 4, s["gender"])
            ws.cell(r, 5, s["dob"])
            ws.cell(r, 6, s["email"])
            ws.cell(r, 7, s["phone"])
            ws.cell(r, 8, s["address"])
            ws.cell(r, 9, s.get("course_name", "N/A"))
            ws.cell(r, 10, s["admission_date"])
            ws.cell(r, 11, s["status"])

        self._autosize(ws)
        wb.save(path)
        messagebox.showinfo("Success", "Students exported successfully")

    # ================= COURSES =================

    def export_courses(self):
        courses = self.app.db.get_all_courses()
        if not courses:
            messagebox.showwarning("No Data", "No courses to export")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"courses_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Courses"

        headers = [
            "Course ID", "Name", "Code", "Description",
            "Duration (Months)", "Fees", "Students Enrolled"
        ]

        self._write_headers(ws, headers, "14a085")

        for r, c in enumerate(courses, 2):
            ws.cell(r, 1, c["course_id"])
            ws.cell(r, 2, c["course_name"])
            ws.cell(r, 3, c["course_code"])
            ws.cell(r, 4, c["description"])
            ws.cell(r, 5, c["duration_months"])
            ws.cell(r, 6, c["fees"])
            ws.cell(r, 7, c["student_count"])

        self._autosize(ws)
        wb.save(path)
        messagebox.showinfo("Success", "Courses exported successfully")

    # ================= ATTENDANCE =================

    def export_attendance_dialog(self):
        dialog = ctk.CTkToplevel(self.app)
        dialog.title("Export Attendance")
        dialog.geometry("400x280")
        dialog.grab_set()

        ctk.CTkLabel(
            dialog,
            text="Export Attendance Records",
            font=ctk.CTkFont(size=18, weight="bold")
        ).pack(pady=20)

        start = DateEntry(dialog, date_pattern="yyyy-mm-dd")
        end = DateEntry(dialog, date_pattern="yyyy-mm-dd")

        start.pack(pady=5)
        end.pack(pady=5)

        def export():
            data = self.app.db.get_attendance_full_by_date_range(start.get_date(), end.get_date())
            if not data:
                messagebox.showwarning("No Data", "No records found")
                return

            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"attendance_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
            )
            if not path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance"

            headers = ["Student ID", "Student Name", "Course", "Date", "Status", "Remarks"]
            self._write_headers(ws, headers, "f39c12")

            for r, rec in enumerate(data, 2):
                ws.cell(r, 1, rec["student_id"])
                ws.cell(r, 2, rec["student_name"])
                ws.cell(r, 3, rec["course_name"])
                ws.cell(r, 4, rec["attendance_date"])
                ws.cell(r, 5, rec["status"])
                ws.cell(r, 6, rec["remarks"])

            self._autosize(ws)
            wb.save(path)
            messagebox.showinfo("Success", "Attendance exported successfully")
            dialog.destroy()

        ctk.CTkButton(dialog, text="Export", command=export).pack(pady=15)

    # ================= HELPERS =================

    def _write_headers(self, ws, headers, color):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    def _autosize(self, ws):
        for col in ws.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = length + 2
